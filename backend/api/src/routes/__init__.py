from flask import Blueprint, jsonify, send_file, request
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import datetime
import requests
from .test import testRoute
from .auth import authRoutes
from config import Config
from models.entities.members import Members
from models.entities.raid import Raid
from models.warClansModels import ModelWarOfClans
from models.entities.warOfClans import WarOfClans
from models.entities.member import WarMember, Member
from models.raidModel import ModelRaid





from services.ClanManager import memberClans
# Crear un blueprint  y registrar las rutas
RaizBlueprint = Blueprint('Raiz', __name__)



RaizBlueprint.register_blueprint(testRoute, url_prefix='/test')
RaizBlueprint.register_blueprint(authRoutes, url_prefix='/auth')


@RaizBlueprint.route('/', methods=['GET'])
def Raiz():
    return "Hello Word"


@RaizBlueprint.route('/clan', methods=['GET'])
def Clan():

    headers = {
    "Authorization": f"Bearer petjw29z"
    }
    ruta = Config.URL_COC+'clans/'+'%23/'+Config.ClanId
    response = requests.get(ruta, headers=headers)
    if response.status_code != 200:
        return jsonify({'error': 'Clan not found'}), 404
    return response.json()






@RaizBlueprint.route('/members', methods=['GET'])
def MembersEndpoint():
    members = memberClans.get_members()

    return jsonify({'members': members.getdict(notNull=True)}), 200

@RaizBlueprint.route('/raids', methods=['GET'])
def Raids():
    mymembers = memberClans.get_members()
    mymembers = memberClans.getRaids(mymembers, AmountRaids=3)

    
    return jsonify({'raids': mymembers.getInfoRaid()}), 200




@RaizBlueprint.route('/test34', methods=['GET'])
def test34():
    r=memberClans.refreshAllClanInfo()
    return jsonify({'message': 'OK'}), 200

@RaizBlueprint.route('/reload', methods=['GET'])
def refreshD():
    try:
        r=memberClans.refreshAllClanInfo()
        return jsonify({'message': 'OK'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@RaizBlueprint.route('/test35', methods=['GET'])
def test35():
    r=ModelRaid.getRaids()
    return jsonify({'message': 'OK', }), 200


lightGrayFill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
darkGrayFill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")

colors = ["00FF00", "80FF80", "FFFF00", "FF8000", "FF0000"]

color_fills = [
    PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Verde (muy bueno)
    PatternFill(start_color="80FF80", end_color="80FF80", fill_type="solid"),  # Verde claro (bueno)
    PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Amarillo (neutral)
    PatternFill(start_color="FF8000", end_color="FF8000", fill_type="solid"),  # Naranja (malo)
    PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   # Rojo (supermalo)
]

#Este endpoin generara un informe de los miembros del clan 
@RaizBlueprint.route('/report', methods=['GET'])
def membersReport():

    template_path = Config.patchTempReport
    wb = load_workbook(template_path)
    ws = wb.active
    columMembers ='B'
    colCreatedAt='C'
    colDonations='R'
    colPedidas='S'
    initMembers=7
    myMembers = memberClans.getAllClanInfo(AmountWars=4)
    ws['P2'] = datetime.datetime.now().strftime('%Y-%m-%d  %H:%M:%S')
    #Iterar los miembros 
    member:Member
    for member in myMembers:
        ws[f'{columMembers}{initMembers+int(member.ranking) -1 }'] = member.username
        ws[f'{colCreatedAt}{initMembers+int(member.ranking) -1 }'] = member.created_at.strftime('%Y-%m-%d')
        ws[f'A{initMembers+int(member.ranking) -1 }'] = member.ranking
        

        cellDonadas = ws[f'{colDonations}{initMembers+int(member.ranking) -1 }']
        cellPedidas = ws[f'{colPedidas}{initMembers+int(member.ranking) -1 }']
        cellDonadas.value = member.donations
        cellPedidas.value = member.troops_requested


        # Calcular la proporción
        if member.troops_requested == 0 and member.donations == 0:
            proportion = 0.2
        elif member.troops_requested == 0: # Evitar división por cero
            proportion = float('inf')  
        else:
            proportion = member.donations /  member.troops_requested

        if proportion > 1.3:  
            fill = color_fills[0]  # Verde (muy bueno)
        elif proportion >= 1:
            fill = color_fills[1]  # Verde claro (bueno)
        elif proportion > 0.4:
            fill = color_fills[2]  # Amarillo (neutral)
        elif proportion > 0.20:
            fill = color_fills[3]  # Naranja (malo)
        else:
            fill = color_fills[4]  # Rojo (supermalo)
        cellPedidas.fill = fill
        cellDonadas.fill = fill


        # Agregar las columnas de Asaltos(ataques y puntos)
        i=1
        colRaids=['D','E','F','G','H','I']
        
        while len(myMembers.raids) >= i:
            print('dentro del while xdd')
            if i >3: break
            raid:Raid
            raid = myMembers.raids[-i]
            i_Asalto = i*2-2   # 0 2 4 
            ws[f'{colRaids[i_Asalto]}5'] = f'Fecha: {raid.startTime.strftime("%Y-%m-%d")}'

            cellAttack = ws[f'{colRaids[i_Asalto]}{initMembers+int(member.ranking) -1 }']
            cellPoints = ws[f'{colRaids[i_Asalto+1]}{initMembers+int(member.ranking) -1 }']
            if raid.isMemberInRaid(member):
                raidInfoMember = next((m for m in raid.members if m == member), None)
                if raidInfoMember:
                    cellAttack.value = f'{raidInfoMember.attacks} / {raidInfoMember.attackLimit}'
                    cellPoints.value = raidInfoMember.resourcesLooted
                    cellPoints.fill = color_fills[1]
                    cellAttack.fill = color_fills[1]
                else:
                    cellAttack.value = '0'
                    cellPoints.value = '0'
                    cellAttack.fill = color_fills[4]
                    cellPoints.fill = color_fills[4]
                    # el miembro Posiblemente estubo Pero No Ataco 
            else:
                # EL miembro No Cumple los requisitos para participár
                pass

            i+=1


        #agregar las coolumnas de Wars (ataque y estreLLas)   
        i=1
        colwars=['J','K','L','M','N','O', 'P','Q']
        while len(myMembers.wars) >= i:
            print('dentro del while xdd')
            if i >4: break
            war:WarOfClans
            war = myMembers.wars[-i]
            i_war=i*2-2
            ws[f'{colwars[i_war]}5'] = f'Fecha: {war.startTime.strftime("%Y-%m-%d")}'
            cellAttack = ws[f'{colwars[i_war]}{initMembers+int(member.ranking) -1 }']
            cellstars = ws[f'{colwars[i_war+1]}{initMembers+int(member.ranking) -1 }']
            
            warinfoMember : WarMember
            warinfoMember= next((m for m in war.members if m == member), None)
            if warinfoMember:
                """ print(f'\n\n {member.username}') """
                cantidadAttack = warinfoMember.lenAttacks()
                cellAttack.value = f'{cantidadAttack} / 2'
                try:
                    if not war.state == 'preparation': 
                        if cantidadAttack ==0: cantidadAttack =-1
                        color= color_fills[3-cantidadAttack]
                        cellAttack.fill = cellstars.fill=color
                except ValueError:
                    pass
                cellstars.value = f'{warinfoMember.getAllStars()} / 6'

            else:
                cellAttack.fill = cellstars.fill = lightGrayFill
                    
                


            i+=1



    # Guardar el archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        output,
        as_attachment=True,
        download_name='reporte_generado.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
